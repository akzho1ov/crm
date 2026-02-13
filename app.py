from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import openpyxl
from openpyxl import load_workbook
import os
import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = 'crm_secret_key_2026_pro'

# Пути к файлам
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'data', 'crm_master.xlsx')

# Хранилище имен менеджеров
MANAGER_NAMES = {i: f'Менеджер {i}' for i in range(1, 21)}

# Пользователи
USERS = {
    'admin': {'password': 'admin123', 'role': 'admin', 'num': 0, 'name': 'Администратор'},
}

# Добавляем менеджеров
for i in range(1, 21):
    USERS[f'manager{i}'] = {
        'password': f'pass{i}',
        'role': 'manager',
        'num': i,
        'name': MANAGER_NAMES[i]
    }

# Маппинг метрик
METRICS = {
    4: 'leads',
    5: 'process',
    6: 'no_answer',
    7: 'waiting',
    8: 'reject',
    9: 'sales',
    10: 'revenue',
    12: 'balance'
}

# ==================== УТИЛИТЫ ====================

def to_int(value):
    """Безопасное преобразование в число"""
    try:
        if value is None or value == '':
            return 0
        if isinstance(value, str) and value.startswith('='):
            return 0
        return int(float(str(value).replace(' ', '').replace(',', '.')))
    except:
        return 0

def to_date_str(value):
    """Преобразование даты в строку"""
    if value is None:
        return ''
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d')
    s = str(value)[:10]
    return s if s != 'None' else ''

# ==================== РАБОТА С EXCEL ====================

def read_manager_data(num):
    """Читает все данные менеджера из Excel"""
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True, read_only=False)
        sheet_name = f'Mgr_{num}'
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return create_empty_manager_data()
        
        sheet = wb[sheet_name]
        
        # Читаем дневные метрики (дни 1-31)
        days = {}
        for day in range(1, 32):
            day_data = {}
            for row, metric_name in METRICS.items():
                cell_value = sheet.cell(row=row, column=day + 1).value
                day_data[metric_name] = to_int(cell_value)
            days[day] = day_data
        
        # Вычисляем итоги
        totals = {metric: 0 for metric in METRICS.values()}
        for day in range(1, 32):
            for metric in METRICS.values():
                totals[metric] += days[day][metric]
        
        # Конверсия и средний чек
        totals['conv'] = round(totals['sales'] / totals['leads'] * 100, 1) if totals['leads'] > 0 else 0
        totals['avg_check'] = round(totals['revenue'] / totals['sales']) if totals['sales'] > 0 else 0
        
        # Читаем клиентов (строки 17-136)
        clients = []
        for row in range(17, 137):
            name = sheet.cell(row=row, column=2).value
            region = sheet.cell(row=row, column=3).value
            phone = sheet.cell(row=row, column=4).value
            
            if name or region or phone:
                clients.append({
                    'row': row,
                    'name': str(name or ''),
                    'region': str(region or ''),
                    'phone': str(phone or ''),
                    'price': to_int(sheet.cell(row=row, column=5).value),
                    'status': str(sheet.cell(row=row, column=6).value or ''),
                    'next_date': to_date_str(sheet.cell(row=row, column=7).value),
                    'balance': to_int(sheet.cell(row=row, column=8).value),
                    'comment': str(sheet.cell(row=row, column=9).value or '')
                })
        
        wb.close()
        
        return {
            'days': days,
            'totals': totals,
            'clients': clients
        }
        
    except Exception as e:
        print(f"Ошибка чтения данных менеджера {num}: {e}")
        return create_empty_manager_data()

def create_empty_manager_data():
    """Создает пустую структуру данных"""
    days = {}
    for day in range(1, 32):
        days[day] = {metric: 0 for metric in METRICS.values()}
    
    totals = {metric: 0 for metric in METRICS.values()}
    totals['conv'] = 0
    totals['avg_check'] = 0
    
    return {
        'days': days,
        'totals': totals,
        'clients': []
    }

def save_daily_data(num, day, data):
    """Сохраняет дневные данные"""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb[f'Mgr_{num}']
        
        for row, metric in METRICS.items():
            value = to_int(data.get(metric, 0))
            sheet.cell(row=row, column=day + 1).value = value if value > 0 else None
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        print(f"Ошибка сохранения: {e}")
        return False

def save_client_data(num, row, data):
    """Сохраняет данные клиента"""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb[f'Mgr_{num}']
        
        sheet.cell(row=row, column=2).value = data.get('name', '')
        sheet.cell(row=row, column=3).value = data.get('region', '')
        sheet.cell(row=row, column=4).value = data.get('phone', '')
        
        price = to_int(data.get('price', 0))
        sheet.cell(row=row, column=5).value = price if price > 0 else None
        
        status = data.get('status', '')
        sheet.cell(row=row, column=6).value = status
        
        next_date = data.get('next_date', '')
        sheet.cell(row=row, column=7).value = next_date if next_date else None
        
        balance = 0 if status == 'Полностью' else price
        sheet.cell(row=row, column=8).value = balance if balance > 0 else None
        
        sheet.cell(row=row, column=9).value = data.get('comment', '')
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        print(f"Ошибка сохранения клиента: {e}")
        return False

def delete_client_data(num, row):
    """Удаляет данные клиента"""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb[f'Mgr_{num}']
        
        for col in range(2, 10):
            sheet.cell(row=row, column=col).value = None
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        print(f"Ошибка удаления клиента: {e}")
        return False

def read_all_managers(period='month', day=None):
    """Читает данные всех менеджеров для дашборда с учетом периода"""
    managers = []
    
    for i in range(1, 21):
        data = read_manager_data(i)
        
        if period == 'day':
            # Данные за один день
            if day is None:
                day = datetime.date.today().day
            day_metrics = data['days'].get(day, {metric: 0 for metric in METRICS.values()})
            
            managers.append({
                'num': i,
                'name': MANAGER_NAMES[i],
                'leads': day_metrics['leads'],
                'process': day_metrics['process'],
                'no_answer': day_metrics['no_answer'],
                'waiting': day_metrics['waiting'],
                'reject': day_metrics['reject'],
                'sales': day_metrics['sales'],
                'conv': round(day_metrics['sales'] / day_metrics['leads'] * 100, 1) if day_metrics['leads'] > 0 else 0,
                'revenue': day_metrics['revenue'],
                'avg_check': round(day_metrics['revenue'] / day_metrics['sales']) if day_metrics['sales'] > 0 else 0,
                'balance': day_metrics['balance'],
                'clients': len(data['clients'])
            })
            
        elif period == 'week':
            # Данные за неделю
            if day is None:
                day = datetime.date.today().day
            start_day = max(1, day - 6)
            
            week_metrics = {metric: 0 for metric in METRICS.values()}
            for d in range(start_day, day + 1):
                for metric in METRICS.values():
                    week_metrics[metric] += data['days'][d][metric]
            
            managers.append({
                'num': i,
                'name': MANAGER_NAMES[i],
                'leads': week_metrics['leads'],
                'process': week_metrics['process'],
                'no_answer': week_metrics['no_answer'],
                'waiting': week_metrics['waiting'],
                'reject': week_metrics['reject'],
                'sales': week_metrics['sales'],
                'conv': round(week_metrics['sales'] / week_metrics['leads'] * 100, 1) if week_metrics['leads'] > 0 else 0,
                'revenue': week_metrics['revenue'],
                'avg_check': round(week_metrics['revenue'] / week_metrics['sales']) if week_metrics['sales'] > 0 else 0,
                'balance': week_metrics['balance'],
                'clients': len(data['clients'])
            })
            
        else:  # month
            managers.append({
                'num': i,
                'name': MANAGER_NAMES[i],
                'leads': data['totals']['leads'],
                'process': data['totals']['process'],
                'no_answer': data['totals']['no_answer'],
                'waiting': data['totals']['waiting'],
                'reject': data['totals']['reject'],
                'sales': data['totals']['sales'],
                'conv': data['totals']['conv'],
                'revenue': data['totals']['revenue'],
                'avg_check': data['totals']['avg_check'],
                'balance': data['totals']['balance'],
                'clients': len(data['clients'])
            })
    
    # Сортируем по выручке
    managers.sort(key=lambda x: x['revenue'], reverse=True)
    
    # Добавляем ранги
    for rank, mgr in enumerate(managers, 1):
        mgr['rank'] = rank
    
    # Итоги
    totals = {
        'leads': sum(m['leads'] for m in managers),
        'process': sum(m['process'] for m in managers),
        'no_answer': sum(m['no_answer'] for m in managers),
        'waiting': sum(m['waiting'] for m in managers),
        'reject': sum(m['reject'] for m in managers),
        'sales': sum(m['sales'] for m in managers),
        'revenue': sum(m['revenue'] for m in managers),
        'balance': sum(m['balance'] for m in managers),
        'clients': sum(m['clients'] for m in managers)
    }
    totals['conv'] = round(totals['sales'] / totals['leads'] * 100, 1) if totals['leads'] > 0 else 0
    totals['avg_check'] = round(totals['revenue'] / totals['sales']) if totals['sales'] > 0 else 0
    
    return {
        'managers': managers,
        'totals': totals
    }

# ==================== ДЕКОРАТОРЫ ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session['user']['role'] != 'admin':
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== РОУТЫ ====================

@app.route('/')
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    if session['user']['role'] == 'admin':
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('manager_page'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if username in USERS and USERS[username]['password'] == password:
            session['user'] = {
                'username': username,
                'role': USERS[username]['role'],
                'num': USERS[username]['num'],
                'name': USERS[username]['name']
            }
            return redirect(url_for('index'))
        else:
            error = 'Неверный логин или пароль'
    
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@admin_required
def dashboard():
    data = read_all_managers()
    return render_template('dashboard.html', data=data, user=session['user'])

@app.route('/manager')
@login_required
def manager_page():
    if session['user']['role'] == 'admin':
        return redirect(url_for('dashboard'))
    
    num = session['user']['num']
    data = read_manager_data(num)
    today = datetime.date.today().day
    
    return render_template('manager.html', 
                         data=data, 
                         user=session['user'],
                         today=today)

# ==================== API ====================

@app.route('/api/save_daily', methods=['POST'])
@login_required
def api_save_daily():
    """Сохраняет дневные метрики"""
    if session['user']['role'] == 'admin':
        return jsonify({'success': False, 'error': 'Доступ запрещен'})
    
    num = session['user']['num']
    day = int(request.form.get('day', 1))
    
    data = {}
    for metric in METRICS.values():
        data[metric] = request.form.get(metric, '0')
    
    if save_daily_data(num, day, data):
        updated_data = read_manager_data(num)
        return jsonify({
            'success': True,
            'totals': updated_data['totals']
        })
    else:
        return jsonify({'success': False, 'error': 'Ошибка сохранения'})

@app.route('/api/period_data', methods=['POST'])
@login_required
def api_period_data():
    """Возвращает данные за период"""
    if session['user']['role'] == 'admin':
        return jsonify({'success': False})
    
    num = session['user']['num']
    period = request.form.get('period', 'day')
    day = int(request.form.get('day', datetime.date.today().day))
    
    data = read_manager_data(num)
    
    if period == 'day':
        metrics = data['days'][day]
        return jsonify({'success': True, 'metrics': metrics})
    
    elif period == 'week':
        start_day = max(1, day - 6)
        metrics = {metric: 0 for metric in METRICS.values()}
        
        for d in range(start_day, day + 1):
            for metric in METRICS.values():
                metrics[metric] += data['days'][d][metric]
        
        return jsonify({'success': True, 'metrics': metrics, 'start_day': start_day, 'end_day': day})
    
    elif period == 'month':
        return jsonify({'success': True, 'metrics': data['totals']})
    
    return jsonify({'success': False})

@app.route('/api/dashboard_period', methods=['POST'])
@admin_required
def api_dashboard_period():
    """Возвращает данные дашборда за период"""
    period = request.form.get('period', 'month')
    day = request.form.get('day')
    
    if day:
        day = int(day)
    
    data = read_all_managers(period=period, day=day)
    return jsonify({
        'success': True,
        'managers': data['managers'],
        'totals': data['totals']
    })

@app.route('/api/save_client', methods=['POST'])
@login_required
def api_save_client():
    """Сохраняет или добавляет клиента"""
    if session['user']['role'] == 'admin':
        return jsonify({'success': False})
    
    num = session['user']['num']
    row = request.form.get('row')
    
    client_data = {
        'name': request.form.get('name', ''),
        'region': request.form.get('region', ''),
        'phone': request.form.get('phone', ''),
        'price': request.form.get('price', '0'),
        'status': request.form.get('status', ''),
        'next_date': request.form.get('next_date', ''),
        'comment': request.form.get('comment', '')
    }
    
    if row:
        row = int(row)
    else:
        current_data = read_manager_data(num)
        used_rows = {c['row'] for c in current_data['clients']}
        row = None
        for r in range(17, 137):
            if r not in used_rows:
                row = r
                break
        
        if row is None:
            return jsonify({'success': False, 'error': 'Нет свободных мест'})
    
    if save_client_data(num, row, client_data):
        updated_data = read_manager_data(num)
        return jsonify({'success': True, 'clients': updated_data['clients']})
    else:
        return jsonify({'success': False, 'error': 'Ошибка сохранения'})

@app.route('/api/delete_client', methods=['POST'])
@login_required
def api_delete_client():
    """Удаляет клиента"""
    if session['user']['role'] == 'admin':
        return jsonify({'success': False})
    
    num = session['user']['num']
    row = int(request.form.get('row'))
    
    if delete_client_data(num, row):
        updated_data = read_manager_data(num)
        return jsonify({'success': True, 'clients': updated_data['clients']})
    else:
        return jsonify({'success': False, 'error': 'Ошибка удаления'})

@app.route('/api/rename_manager', methods=['POST'])
@admin_required
def api_rename_manager():
    """Переименовывает менеджера"""
    num = int(request.form.get('num'))
    new_name = request.form.get('name', '').strip()
    
    if 1 <= num <= 20 and new_name:
        MANAGER_NAMES[num] = new_name
        USERS[f'manager{num}']['name'] = new_name
        return jsonify({'success': True})
    
    return jsonify({'success': False})

if __name__ == '__main__':
    print('='*60)
    print('🚀 CRM СИСТЕМА v4.1 PROFESSIONAL')
    print('='*60)
    print(f'📂 Excel файл: {EXCEL_FILE}')
    print('🌐 Открыть: http://localhost:5050')
    print('='*60)
    print('👤 Админ: admin / admin123')
    print('👤 Менеджер: manager1 / pass1')
    print('='*60)
    
    app.run(host='0.0.0.0', port=5050, debug=True)
